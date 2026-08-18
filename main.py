from pathlib import Path
from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import RedirectResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import database as db
import json, io, uuid
from urllib.parse import quote
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
app = FastAPI(title="신입사원 온보딩 자동화 시스템")
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")
db.init_db()
db.ensure_phase3_schema()
db.ensure_v12_schema()

ROLE_CODES = {"hr": "인사담당자", "manager": "부서장", "newhire": "신입사원"}
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
ALLOWED_UPLOAD_EXTENSIONS = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".jpg", ".jpeg", ".png"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024


def current_role(request: Request):
    return ROLE_CODES.get(request.cookies.get("role_code", "hr"), "인사담당자")


def current_employee_id(request: Request):
    raw = request.cookies.get("employee_id", "")
    try:
        return int(raw) if raw else None
    except ValueError:
        return None


def perm(request: Request):
    return db.get_permission(current_role(request))


def can_view_employee(request: Request, employee_id: int):
    p = perm(request)
    if p and p["can_view_all"]:
        return True
    return current_role(request) == "신입사원" and current_employee_id(request) == employee_id


def context(request: Request, **kwargs):
    data = {
        "request": request,
        "settings": db.get_settings(),
        "current_role": current_role(request),
        "current_employee_id": current_employee_id(request),
        "permission": perm(request),
        "modules": db.list_modules(False),
    }
    data.update(kwargs)
    return data


async def save_upload(file: UploadFile | None):
    if not file or not file.filename:
        return "", ""
    suffix = Path(file.filename).suffix.lower()
    if suffix not in ALLOWED_UPLOAD_EXTENSIONS:
        raise ValueError("허용되지 않은 파일 형식입니다. PDF, Office 문서, 이미지 파일만 업로드할 수 있습니다.")
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("파일 크기는 10MB를 초과할 수 없습니다.")
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    (UPLOAD_DIR / stored_name).write_bytes(content)
    return file.filename, stored_name


@app.get("/switch-role/{code}")
def switch_role(code: str):
    if code not in ROLE_CODES:
        code = "hr"
    target = "/me" if code == "newhire" else "/"
    response = RedirectResponse(target, status_code=303)
    response.set_cookie("role_code", code, httponly=True, samesite="lax")
    if code != "newhire":
        response.delete_cookie("employee_id")
    return response


@app.get("/employees/{employee_id}/preview-newhire")
def preview_newhire(request: Request, employee_id: int):
    if not perm(request) or not perm(request)["can_view_all"]:
        return RedirectResponse("/", 303)
    if not db.get_employee(employee_id):
        return RedirectResponse("/employees", 303)
    response = RedirectResponse("/me", 303)
    response.set_cookie("role_code", "newhire", httponly=True, samesite="lax")
    response.set_cookie("employee_id", str(employee_id), httponly=True, samesite="lax")
    return response


@app.get("/exit-preview")
def exit_preview():
    response = RedirectResponse("/", 303)
    response.set_cookie("role_code", "hr", httponly=True, samesite="lax")
    response.delete_cookie("employee_id")
    return response


@app.get("/me")
def my_onboarding(request: Request):
    employee_id = current_employee_id(request)
    if current_role(request) != "신입사원" or not employee_id:
        return templates.TemplateResponse(
            request=request,
            name="my_onboarding_empty.html",
            context=context(request),
        )
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.get("/")
def dashboard(request: Request):
    if current_role(request) == "신입사원":
        return RedirectResponse("/me", 303)
    db.refresh_notifications()
    employees = [e for e in db.list_employees() if e["status"] != "입사취소"][:8] if (perm(request) and perm(request)["can_view_all"]) else []
    rows = [{"employee": e, "progress": db.employee_progress(e["id"])} for e in employees]
    return templates.TemplateResponse(
        request=request,
        name="dashboard.html",
        context=context(
            request,
            summary=db.dashboard_summary(),
            employee_rows=rows,
            notifications=(db.list_notifications()[:5] if (perm(request) and perm(request)["can_view_all"]) else []),
        ),
    )


@app.get("/employees")
def employees(request: Request, q: str = "", department: str = "", status: str = ""):
    if not perm(request) or not perm(request)["can_view_all"]:
        return RedirectResponse("/", 303)
    rows = db.list_employees(q, department, status)
    enriched = [{"employee": r, "progress": db.employee_progress(r["id"])} for r in rows]
    departments = sorted({r["department"] for r in db.list_employees()})
    return templates.TemplateResponse(
        request=request,
        name="employees.html",
        context=context(request, rows=enriched, q=q, department=department, status=status, departments=departments),
    )


@app.get("/employees/new")
def employee_new(request: Request):
    if not perm(request) or not perm(request)["can_edit_employee"]:
        return RedirectResponse("/employees", 303)
    return templates.TemplateResponse(
        request=request,
        name="employee_form.html",
        context=context(request, error=request.query_params.get("error", "")),
    )


@app.post("/employees/new")
def employee_create(
    request: Request,
    name: str = Form(...),
    employee_no: str = Form(""),
    department: str = Form(...),
    job_title: str = Form(""),
    employment_type: str = Form("정규직"),
    start_date: str = Form(...),
):
    if not perm(request) or not perm(request)["can_edit_employee"]:
        return RedirectResponse("/employees", 303)
    try:
        employee_id = db.create_employee(name, employee_no, department, job_title, employment_type, start_date)
    except (ValueError, TypeError) as exc:
        return RedirectResponse(f"/employees/new?error={quote(str(exc))}", 303)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.get("/employees/{employee_id}")
def employee_detail(request: Request, employee_id: int):
    if not can_view_employee(request, employee_id):
        return RedirectResponse("/", 303)
    employee = db.get_employee(employee_id)
    if not employee:
        return RedirectResponse("/employees", 303)
    return templates.TemplateResponse(
        request=request,
        name="employee_detail.html",
        context=context(
            request,
            employee=employee,
            tasks=db.list_tasks(employee_id),
            progress=db.employee_progress(employee_id),
            employee_documents=db.list_employee_documents(employee_id),
            sync_message=request.query_params.get("sync", ""),
            error=request.query_params.get("error", ""),
        ),
    )


@app.post("/tasks/{task_id}/toggle")
def task_toggle(request: Request, task_id: int, employee_id: int = Form(...)):
    if perm(request) and perm(request)["can_edit_employee"]:
        db.toggle_task(task_id)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.get("/tasks/{task_id}/edit")
def task_edit_page(request: Request, task_id: int):
    if not perm(request) or not perm(request)["can_edit_employee"]:
        return RedirectResponse("/", 303)
    task = db.get_task(task_id)
    if not task:
        return RedirectResponse("/employees", 303)
    employee = db.get_employee(task["employee_id"])
    return templates.TemplateResponse(request=request, name="task_edit.html", context=context(request, task=task, employee=employee))


@app.post("/tasks/{task_id}/edit")
def task_edit_save(
    request: Request,
    task_id: int,
    employee_id: int = Form(...),
    title: str = Form(...),
    owner_department: str = Form(...),
    assigned_to: str = Form(""),
    due_date: str = Form(...),
    required: int = Form(0),
    notes: str = Form(""),
):
    if perm(request) and perm(request)["can_edit_employee"]:
        try:
            db.update_task(task_id, title, owner_department, assigned_to, due_date, required, notes)
        except ValueError as exc:
            return RedirectResponse(f"/employees/{employee_id}?error={quote(str(exc))}", 303)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/tasks/{task_id}/cancel")
def task_cancel(request: Request, task_id: int, employee_id: int = Form(...), cancelled: int = Form(1)):
    if perm(request) and perm(request)["can_edit_employee"]:
        db.set_task_cancelled(task_id, bool(cancelled))
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employees/{employee_id}/tasks/new")
def employee_task_new(
    request: Request,
    employee_id: int,
    title: str = Form(...),
    owner_department: str = Form(...),
    assigned_to: str = Form(""),
    due_date: str = Form(...),
    required: int = Form(0),
    notes: str = Form(""),
):
    if perm(request) and perm(request)["can_edit_employee"]:
        try:
            db.add_employee_task(employee_id, title, owner_department, assigned_to, due_date, required, notes)
        except ValueError as exc:
            return RedirectResponse(f"/employees/{employee_id}?error={quote(str(exc))}", 303)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employees/{employee_id}/sync-template")
def employee_sync_template(request: Request, employee_id: int):
    if not perm(request) or not perm(request)["can_edit_employee"]:
        return RedirectResponse(f"/employees/{employee_id}", 303)
    result = db.sync_employee_from_template(employee_id)
    msg = f"{result['template']} · 추가 {result['added']}건 · 수정 {result['updated']}건 · 완료/취소 보존 {result['preserved']}건"
    return RedirectResponse(f"/employees/{employee_id}?sync={quote(msg)}", 303)


@app.post("/employees/{employee_id}/info")
def employee_info_update(
    request: Request,
    employee_id: int,
    name: str = Form(...),
    employee_no: str = Form(""),
    department: str = Form(...),
    job_title: str = Form(""),
    employment_type: str = Form("정규직"),
    reapply_template: int = Form(0),
):
    if not perm(request) or not perm(request)["can_edit_employee"]:
        return RedirectResponse(f"/employees/{employee_id}", 303)
    try:
        result = db.update_employee_info(employee_id, name, employee_no, department, job_title, employment_type, bool(reapply_template))
    except ValueError as exc:
        return RedirectResponse(f"/employees/{employee_id}?error={quote(str(exc))}", 303)
    if result:
        msg = f"기본정보 수정 + 템플릿 재선택 · 추가 {result['added']}건 · 수정 {result['updated']}건 · 보존 {result['preserved']}건"
        return RedirectResponse(f"/employees/{employee_id}?sync={quote(msg)}", 303)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employees/{employee_id}/exception")
def employee_exception(request: Request, employee_id: int, exception_status: str = Form("")):
    if perm(request) and perm(request)["can_edit_employee"]:
        db.set_exception(employee_id, exception_status)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employees/{employee_id}/start-date")
def employee_start_date_change(request: Request, employee_id: int, new_start_date: str = Form(...), shift_tasks: int = Form(0)):
    if perm(request) and perm(request)["can_edit_employee"]:
        db.change_employee_start_date(employee_id, new_start_date, bool(shift_tasks))
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employees/{employee_id}/cancel")
def employee_cancel(request: Request, employee_id: int):
    if perm(request) and perm(request)["can_edit_employee"]:
        db.cancel_employee(employee_id)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employee-documents/{employee_document_id}/upload")
async def employee_document_upload(request: Request, employee_document_id: int, employee_id: int = Form(...), file: UploadFile = File(...)):
    if not can_view_employee(request, employee_id):
        return RedirectResponse("/", 303)
    # 인사담당자 또는 본인 미리보기만 업로드 가능
    if not ((perm(request) and perm(request)["can_edit_employee"]) or (current_role(request) == "신입사원" and current_employee_id(request) == employee_id)):
        return RedirectResponse(f"/employees/{employee_id}", 303)
    row = db.get_employee_document(employee_document_id)
    if not row or row["employee_id"] != employee_id:
        return RedirectResponse(f"/employees/{employee_id}", 303)
    try:
        file_name, stored_name = await save_upload(file)
    except ValueError as exc:
        return RedirectResponse(f"/employees/{employee_id}?error={quote(str(exc))}", 303)
    if row["stored_name"]:
        old = UPLOAD_DIR / row["stored_name"]
        if old.exists():
            old.unlink()
    actor = "신입사원" if current_role(request) == "신입사원" else "인사담당자"
    db.submit_employee_document(employee_document_id, file_name, stored_name, actor)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.post("/employee-documents/{employee_document_id}/review")
def employee_document_review(
    request: Request,
    employee_document_id: int,
    employee_id: int = Form(...),
    status: str = Form(...),
    note: str = Form(""),
):
    if perm(request) and perm(request)["can_edit_employee"]:
        try:
            db.review_employee_document(employee_document_id, status, note)
        except ValueError as exc:
            return RedirectResponse(f"/employees/{employee_id}?error={quote(str(exc))}", 303)
    return RedirectResponse(f"/employees/{employee_id}", 303)


@app.get("/employee-documents/{employee_document_id}/download")
def employee_document_download(request: Request, employee_document_id: int):
    row = db.get_employee_document(employee_document_id)
    if not row or not can_view_employee(request, row["employee_id"]):
        return RedirectResponse("/", 303)
    path = UPLOAD_DIR / row["stored_name"] if row["stored_name"] else None
    if not path or not path.exists():
        return RedirectResponse(f"/employees/{row['employee_id']}", 303)
    return FileResponse(path, filename=row["file_name"])


@app.get("/templates")
def template_list(request: Request):
    if not perm(request) or not perm(request)["can_manage_templates"]:
        return RedirectResponse("/", 303)
    return templates.TemplateResponse(request=request, name="templates.html", context=context(request, templates_data=db.list_templates()))


@app.post("/templates/new")
def template_new(request: Request, name: str = Form(...), match_department: str = Form(""), match_job_title: str = Form(""), match_employment_type: str = Form(""), process_type: str = Form("온보딩")):
    if perm(request) and perm(request)["can_manage_templates"]:
        tid = db.create_template(name, match_department, match_job_title, match_employment_type, process_type)
        return RedirectResponse(f"/templates/{tid}", 303)
    return RedirectResponse("/", 303)


@app.get("/templates/{template_id}")
def template_detail(request: Request, template_id: int):
    if not perm(request) or not perm(request)["can_manage_templates"]:
        return RedirectResponse("/", 303)
    template = db.get_template(template_id)
    if not template:
        return RedirectResponse("/templates", 303)
    return templates.TemplateResponse(request=request, name="template_detail.html", context=context(request, template=template, items=db.list_template_items(template_id)))


@app.post("/templates/{template_id}/update")
def template_update(request: Request, template_id: int, name: str = Form(...), match_department: str = Form(""), match_job_title: str = Form(""), match_employment_type: str = Form(""), active: int = Form(0), process_type: str = Form("온보딩")):
    if perm(request) and perm(request)["can_manage_templates"]:
        db.update_template(template_id, name, match_department, match_job_title, match_employment_type, active, process_type)
    return RedirectResponse(f"/templates/{template_id}", 303)


@app.post("/templates/{template_id}/items/new")
def template_item_new(request: Request, template_id: int, title: str = Form(...), owner_department: str = Form(...), due_offset: int = Form(0), required: int = Form(0), sort_order: int = Form(100)):
    if perm(request) and perm(request)["can_manage_templates"]:
        db.add_template_item(template_id, title, owner_department, due_offset, required, sort_order)
    return RedirectResponse(f"/templates/{template_id}", 303)


@app.post("/templates/{template_id}/items/{item_id}/update")
def template_item_update(request: Request, template_id: int, item_id: int, title: str = Form(...), owner_department: str = Form(...), due_offset: int = Form(0), required: int = Form(0), enabled: int = Form(0), sort_order: int = Form(100)):
    if perm(request) and perm(request)["can_manage_templates"]:
        db.update_template_item(item_id, title, owner_department, due_offset, required, enabled, sort_order)
    return RedirectResponse(f"/templates/{template_id}", 303)


@app.post("/templates/{template_id}/items/{item_id}/delete")
def template_item_delete(request: Request, template_id: int, item_id: int):
    if perm(request) and perm(request)["can_manage_templates"]:
        db.delete_template_item(item_id)
    return RedirectResponse(f"/templates/{template_id}", 303)


@app.get("/notifications")
def notifications(request: Request):
    if not perm(request) or not perm(request)["can_view_all"]:
        return RedirectResponse("/", 303)
    return templates.TemplateResponse(request=request, name="notifications.html", context=context(request, rows=db.list_notifications()))


@app.post("/notifications/{notification_id}/ack")
def notification_ack(request: Request, notification_id: int):
    if perm(request) and perm(request)["can_view_all"]:
        db.acknowledge_notification(notification_id)
    return RedirectResponse("/notifications", 303)


@app.get("/settings")
def settings(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    return templates.TemplateResponse(
        request=request,
        name="settings.html",
        context=context(request, logs=db.list_change_logs(), permissions=db.list_permissions(), owners=db.list_department_owners(), modules_data=db.list_modules()),
    )


@app.post("/settings/names")
def settings_names(
    request: Request,
    program_title: str = Form(...),
    feature_employee_add: str = Form(...),
    feature_progress: str = Form(...),
    feature_overdue: str = Form(...),
    feature_template: str = Form(...),
    feature_permissions: str = Form(...),
):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    values = {
        "program_title": program_title,
        "feature_employee_add": feature_employee_add,
        "feature_progress": feature_progress,
        "feature_overdue": feature_overdue,
        "feature_template": feature_template,
        "feature_permissions": feature_permissions,
    }
    for key, value in values.items():
        db.update_setting(key, value)
    return RedirectResponse("/settings", 303)


@app.post("/settings/alerts")
def settings_alerts(request: Request, notify_before_days: int = Form(2), escalate_after_days: int = Form(3)):
    if perm(request) and perm(request)["can_manage_settings"]:
        db.update_setting("notify_before_days", str(max(0, notify_before_days)))
        db.update_setting("escalate_after_days", str(max(1, escalate_after_days)))
    return RedirectResponse("/settings", 303)


@app.post("/settings/permissions")
def settings_permissions(request: Request, role: str = Form(...), can_view_all: int = Form(0), can_edit_employee: int = Form(0), can_manage_templates: int = Form(0), can_manage_settings: int = Form(0), can_manage_permissions: int = Form(0)):
    if perm(request) and perm(request)["can_manage_permissions"]:
        db.update_permission(role, can_view_all, can_edit_employee, can_manage_templates, can_manage_settings, can_manage_permissions)
    return RedirectResponse("/settings", 303)


@app.post("/settings/owners")
def settings_owners(request: Request, department: str = Form(...), owner_name: str = Form(""), contact: str = Form("")):
    if perm(request) and perm(request)["can_manage_settings"]:
        db.upsert_department_owner(department, owner_name, contact)
    return RedirectResponse("/settings", 303)


@app.post("/settings/modules")
def settings_module_update(request: Request, code: str = Form(...), display_name: str = Form(...), enabled: int = Form(0), sort_order: int = Form(100)):
    if perm(request) and perm(request)["can_manage_settings"]:
        db.update_module(code, display_name, enabled, sort_order)
    return RedirectResponse("/settings", 303)


@app.get("/documents")
def documents(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    return templates.TemplateResponse(request=request, name="documents.html", context=context(request, rows=db.list_document_forms(), error=request.query_params.get("error", "")))


@app.post("/documents/new")
async def document_new(request: Request, name: str = Form(...), category: str = Form("기타"), required: int = Form(0), file: UploadFile | None = File(None)):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    try:
        file_name, stored_name = await save_upload(file)
    except ValueError as exc:
        return RedirectResponse(f"/documents?error={quote(str(exc))}", 303)
    db.add_document_form(name, category, required, file_name, stored_name)
    return RedirectResponse("/documents", 303)


@app.post("/documents/{doc_id}/update")
def document_update(request: Request, doc_id: int, name: str = Form(...), category: str = Form("기타"), required: int = Form(0), active: int = Form(0)):
    if perm(request) and perm(request)["can_manage_settings"]:
        db.update_document_form(doc_id, name, category, required, active)
    return RedirectResponse("/documents", 303)


@app.get("/documents/{doc_id}/download")
def document_download(request: Request, doc_id: int):
    row = db.get_document_form(doc_id)
    if not row or not row["active"] or not row["stored_name"] or not (UPLOAD_DIR / row["stored_name"]).exists():
        return RedirectResponse("/", 303)
    # 활성 양식은 로그인 역할에 관계없이 온보딩 참여자가 다운로드 가능
    if current_role(request) not in ROLE_CODES.values():
        return RedirectResponse("/", 303)
    return FileResponse(UPLOAD_DIR / row["stored_name"], filename=row["file_name"])


@app.post("/documents/{doc_id}/delete")
def document_delete(request: Request, doc_id: int):
    if perm(request) and perm(request)["can_manage_settings"]:
        row = db.delete_document_form(doc_id)
        if row and row["stored_name"]:
            path = UPLOAD_DIR / row["stored_name"]
            if path.exists():
                path.unlink()
    return RedirectResponse("/documents", 303)


@app.get("/data")
def data_page(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    return templates.TemplateResponse(
        request=request,
        name="data.html",
        context=context(
            request,
            imported=request.query_params.get("imported", ""),
            skipped=request.query_params.get("skipped", ""),
            errors=request.query_params.get("errors", ""),
        ),
    )


@app.get("/data/export.xlsx")
def export_excel(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    wb = Workbook(); ws = wb.active; ws.title = "입사자"
    ws.append(["사번", "성명", "소속부서", "직위/직책", "고용형태", "입사일", "상태", "예외상태", "온보딩진행률"])
    for e in db.list_employees():
        ws.append([e["employee_no"], e["name"], e["department"], e["job_title"], e["employment_type"], e["start_date"], e["status"], e["exception_status"], db.employee_progress(e["id"])])
    ws2 = wb.create_sheet("온보딩업무"); ws2.append(["입사자", "소속부서", "업무명", "담당부서", "담당자", "완료기한", "필수여부", "상태", "메모"])
    for e in db.list_employees():
        for t in db.list_tasks(e["id"]):
            ws2.append([e["name"], e["department"], t["title"], t["owner_department"], t["assigned_to"], t["due_date"], "필수" if t["required"] else "권장", t["status"], t["notes"] or ""])
    ws3 = wb.create_sheet("입사서류")
    ws3.append(["입사자", "소속부서", "서류명", "구분", "필수여부", "상태", "제출파일", "확인메모"])
    for e in db.list_employees():
        for d in db.list_employee_documents(e["id"]):
            ws3.append([e["name"], e["department"], d["name"], d["category"], "필수" if d["required"] else "권장", d["status"], d["file_name"], d["note"]])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=onboarding_export.xlsx"})


@app.get("/data/import-template.xlsx")
def import_template_excel(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    wb = Workbook(); ws = wb.active; ws.title = "입사자등록"
    ws.append(["사번", "성명*", "소속부서*", "직위/직책", "고용형태", "입사일*(YYYY-MM-DD)"])
    ws.append(["2026001", "홍길동", "경영지원팀", "사원", "정규직", "2026-09-01"])
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=employee_import_template.xlsx"})


@app.post("/data/import")
async def import_excel(request: Request, file: UploadFile = File(...)):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    content = await file.read()
    count = skipped = errors = 0
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        return RedirectResponse("/data?errors=1", 303)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        employee_no, name, department, job_title, employment_type, start_date = (list(row) + [None] * 6)[:6]
        if not name or not department or not start_date:
            errors += 1
            continue
        if hasattr(start_date, "strftime"):
            start_date = start_date.strftime("%Y-%m-%d")
        try:
            db.create_employee(str(name), str(employee_no or ""), str(department), str(job_title or ""), str(employment_type or "정규직"), str(start_date))
            count += 1
        except ValueError as exc:
            if "중복 입사자" in str(exc):
                skipped += 1
            else:
                errors += 1
        except Exception:
            errors += 1
    db.log_change("인사담당자", "Excel 입사자 일괄등록", "data", f"등록 {count}명 / 중복 제외 {skipped}명 / 오류 {errors}건")
    return RedirectResponse(f"/data?imported={count}&skipped={skipped}&errors={errors}", 303)


@app.get("/data/database-backup.db")
def database_backup(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    payload = db.create_database_backup_bytes()
    return StreamingResponse(io.BytesIO(payload), media_type="application/octet-stream", headers={"Content-Disposition": "attachment; filename=onboarding_database_backup.db"})


@app.get("/settings/backup.json")
def config_backup(request: Request):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    payload = json.dumps(db.export_config(), ensure_ascii=False, indent=2).encode("utf-8")
    return StreamingResponse(io.BytesIO(payload), media_type="application/json", headers={"Content-Disposition": "attachment; filename=onboarding_settings_backup.json"})


@app.post("/settings/restore")
async def config_restore(request: Request, file: UploadFile = File(...)):
    if not perm(request) or not perm(request)["can_manage_settings"]:
        return RedirectResponse("/", 303)
    try:
        payload = json.loads((await file.read()).decode("utf-8"))
        db.restore_config(payload)
        db.ensure_v12_schema()
    except Exception as exc:
        return RedirectResponse(f"/settings?error={quote(str(exc))}", 303)
    return RedirectResponse("/settings?restored=1", 303)
